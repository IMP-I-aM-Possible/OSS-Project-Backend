import pymysql
import openpyxl as op
conn = pymysql.connect(host='localhost', user='test', password='test', db='oss', charset='utf8')

cur = conn.cursor()
tname = ["linoleicacid","alphalinoleicacid","epadha","methionine","leucine","isoleucine","valine","lysine","phenylalanine","tyrosine","threonine","tryptophan","histamine","vitamina","vitamind","vitamine","vitamink","vitaminc","riboflavin","niacin","vitaminb","pyridoxine","folate","cobalamin","pantothenic","biotin","calcium","phosphorus","natrium","chlorine","potassium","magnesium","iron","zinc","cu","fluorine","mangan","iodine","selenium","molybdenum","chrome"]
lenght = cur.execute("SELECT * FROM "+tname[0])
tname2 = ["pyridoxine","cobalamin"]
a = cur.fetchall()
print(len(tname))
wb = op.load_workbook("nutrient.xlsx")
for i in range(len(tname2)):
    ws = wb[tname2[i]]
    """ ws.cell(row=1,column=5).value="type" """
    lenght = cur.execute("SELECT * FROM "+tname2[i])
    a = cur.fetchall()
    for j in range(lenght):
        commend = str(a[j][3]).replace("μg","")
        commend = commend.replace("g","")
        commend = commend.replace("m","")

        max2 = str(a[j][4]).replace("μg","")
        max2 = max2.replace("g","")
        max2 = max2.replace("m","")
        ws.cell(row=j+2,column=1).value=a[j][1]
        ws.cell(row=j+2,column=2).value=a[j][2]
        ws.cell(row=j+2,column=3).value=commend
        ws.cell(row=j+2,column=4).value=max2

wb.save("nutrient.xlsx")
