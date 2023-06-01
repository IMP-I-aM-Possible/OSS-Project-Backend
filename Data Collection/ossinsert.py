import pandas as pd
from pytest import data
import pymysql
import openpyxl as op
conn = pymysql.connect(host='localhost', user='test', password='test', db='oss', charset='utf8')
cur = conn.cursor()
a = pd.read_excel('osstest2.xlsx',sheet_name="Sheet1",engine='openpyxl')
""" CREATE TABLE `oss`.`nutritional` (
    `nutritional_id` VARCHAR(100) NOT NULL,
    `nutritional_name` TEXT(500) NOT NULL,
    `price` VARCHAR(45) NOT NULL,
    `barcode` VARCHAR(45) NOT NULL,
    `stars` VARCHAR(45) NOT NULL,
    `caution` TEXT(500) NOT NULL,
    `nutritional_info` TEXT(500) NOT NULL,
    `link` TEXT(500) NOT NULL,
    PRIMARY KEY (`nutritional_id`)); """

f1 = a.loc[:,"id"]
f2 = a.loc[:,"name"]
f3 = a.loc[:,"price"]
f4 = a.loc[:,"barcode"]
f5 = a.loc[:,"stars"]
f6 = a.loc[:,"caution"]
f7 = a.loc[:,"info"]
f8 = a.loc[:,"link"]

#1028,1029
wb = op.load_workbook("osstest2.xlsx")
ws = wb["Sheet1"]


    #wb.save("osstest.xlsx") 1028,1029              mcg mg, IU μg
""" for i in range(1028,len(f1)):
    print(ws.cell(row=i+2,column=1).value)
    abcde = ws.cell(row=i+2,column=7).value.replace("mcg"," mcg")
    abcde = abcde.replace("mg"," mg")
    abcde = abcde.replace("IU"," IU")
    abcde = abcde.replace("μg"," μg")
    print(abcde)
    print(f1[i])
    ws.cell(row=i+2,column=7).value = abcde

wb.save("osstest2.xlsx") """
for i in range(1028,len(f1)):
    abcde = f7[i].split("@")[0]
    if abcde !="Serving":
        try:
            cur = conn.cursor()
            body = (f1[i],f2[i],f3[i],f4[i],f5[i],f6[i],f7[i],f8[i],f7[i].split("@")[0])
            cur.execute("INSERT INTO ntest (nutritional_id,nutritional_name,price,barcode,stars,caution,nutritional_info,link,eating) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",body)
            cur.close()
            conn.commit()
        except:
            print(f1[i],i)
""" a = f7[i].split("@")[1]
b = f1[i]
body = (a,b)
cur = conn.cursor()
cur.execute("UPDATE nutritional set eating = %s WHERE nutritional_id = %s",body)
cur.close()
conn.commit() """
