import pandas as pd
import openpyxl as op
from pytest import data
import pymysql
sname = ["linoleicacid","alphalinoleicacid","epa","dha","methionine","leucine","isoleucine","valine","lysine","phenylalanine","tyrosine","threonine","tryptophan","histamine","vitamina","vitamind","vitamine","vitamink","vitaminc","riboflavin","niacin","vitaminb","pyridoxine","folate","cobalamin","pantothenic","biotin","calcium","phosphorus","natrium","chlorine","potassium","magnesium","iron","zinc","cu","fluorine","mangan","iodine","selenium","molybdenum","chrome"]
conn = pymysql.connect(host='localhost', user='test', password='test', db='oss', charset='utf8')
cur = conn.cursor()
wb = op.load_workbook("ossnutrient.xlsx")

""" CREATE TABLE `oss`.`delaynutrient` (
    `idx` INT NOT NULL AUTO_INCREMENT,
    `age` FLOAT NOT NULL,
    `gender` VARCHAR(10) NOT NULL,
    `commend` VARCHAR(20) NULL,
    `max` VARCHAR(20) NULL,
    `unit` VARCHAR(10) NULL,
    `nutrient_name` VARCHAR(45) NULL,
    PRIMARY KEY (`idx`)); """


for i in range(len(sname)):
    ws = wb[sname[i]]
    for j in range(2,149):
        f1 = ws.cell(row=j,column=1).value
        f2 = ws.cell(row=j,column=2).value
        f3 = ws.cell(row=j,column=3).value
        f4 = ws.cell(row=j,column=4).value
        f5 = ws.cell(row=j,column=5).value
        f6 = ws.cell(row=j,column=6).value
        body = (f1,f2,f3,f4,f5,f6)
        cur.execute("INSERT INTO dailynutrient (age,gender,commend,max,unit,nutrient_name) VALUES (%s,%s,%s,%s,%s,%s)",body)
        conn.commit()
