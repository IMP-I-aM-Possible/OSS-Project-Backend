import pymysql
import openpyxl as op
import json
import ast
conn = pymysql.connect(host='localhost', user='test', password='test', db='oss', charset='utf8')

cur = conn.cursor()

def insert():
    lenght = cur.execute("Select * from nutritional")
    a = cur.fetchall()
    
    wb = op.load_workbook("ossexcel2.xlsx")
    #print(lenght)
    ws = wb["Sheet1"]
    """ cnt = 1
    for i in a[0]:
        
        print(i,cnt)
        cnt += 1
    alla = ast.literal_eval(a[0][13])
    print(alla['고유 혼합물 Neprinol AFD']) """
    
    for i in range(lenght):
        ws.cell(row=i+2,column=1).value = a[i][0]
        ws.cell(row=i+2,column=2).value = a[i][1]
        ws.cell(row=i+2,column=3).value = a[i][2]
        ws.cell(row=i+2,column=4).value = a[i][3]
        ws.cell(row=i+2,column=5).value = a[i][4]
        ws.cell(row=i+2,column=6).value = a[i][5]
        ws.cell(row=i+2,column=7).value = a[i][6]
        ws.cell(row=i+2,column=8).value = a[i][7]
        ws.cell(row=i+2,column=9).value = a[i][8]
        ws.cell(row=i+2,column=10).value = a[i][9]
        ws.cell(row=i+2,column=11).value = a[i][10]
        ws.cell(row=i+2,column=12).value = a[i][11]
        ws.cell(row=i+2,column=13).value = a[i][12]
        ws.cell(row=i+2,column=14).value = a[i][13]
    wb.save("ossexcel2.xlsx")
insert()